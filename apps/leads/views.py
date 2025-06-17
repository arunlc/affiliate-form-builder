# apps/leads/views.py - VERSION WITHOUT PANDAS
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta
from .models import Lead, LeadNote
from .serializers import LeadSerializer, LeadNoteSerializer
from apps.affiliates.models import Affiliate

# Use openpyxl directly instead of pandas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import logging

logger = logging.getLogger(__name__)

class LeadViewSet(viewsets.ModelViewSet):
    serializer_class = LeadSerializer
    permission_classes = [IsAuthenticated]
    queryset = Lead.objects.all()
    
    def get_queryset(self):
        user = self.request.user
        
        # Base queryset with optimizations
        queryset = Lead.objects.select_related(
            'form', 'affiliate', 'affiliate__user'
        ).prefetch_related('lead_notes')
        
        # Filter by user role
        if user.user_type == 'admin':
            # Admins see all leads
            queryset = queryset.all()
        elif user.user_type == 'operations':
            # Operations see all leads
            queryset = queryset.all()
        else:
            queryset = Lead.objects.none()
        
        # Apply filters from query parameters
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(email__icontains=search) |
                Q(name__icontains=search) |
                Q(form_data__icontains=search)
            )
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        utm_source = self.request.query_params.get('utm_source')
        if utm_source:
            queryset = queryset.filter(utm_source=utm_source)
        
        affiliate_code = self.request.query_params.get('affiliate')
        if affiliate_code and user.user_type in ['admin', 'operations']:
            queryset = queryset.filter(affiliate__affiliate_code=affiliate_code)
        
        form_id = self.request.query_params.get('form')
        if form_id:
            queryset = queryset.filter(form__id=form_id)
        
        date_range = self.request.query_params.get('date_range')
        if date_range:
            now = timezone.now()
            if date_range == 'today':
                queryset = queryset.filter(created_at__date=now.date())
            elif date_range == 'week':
                week_ago = now - timedelta(days=7)
                queryset = queryset.filter(created_at__gte=week_ago)
            elif date_range == 'month':
                month_ago = now - timedelta(days=30)
                queryset = queryset.filter(created_at__gte=month_ago)
            elif date_range == 'quarter':
                quarter_ago = now - timedelta(days=90)
                queryset = queryset.filter(created_at__gte=quarter_ago)
        
        return queryset.order_by('-created_at')
    
    def update(self, request, *args, **kwargs):
        """Update lead - with affiliate restrictions"""
        instance = self.get_object()
        user = request.user
        
        # Check if affiliate can only update their own leads
        if user.user_type == 'affiliate':
            try:
                affiliate = Affiliate.objects.get(user=user)
                if instance.affiliate != affiliate:
                    return Response(
                        {'error': 'You can only update your own leads'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            except Affiliate.DoesNotExist:
                return Response(
                    {'error': 'Affiliate profile not found'}, 
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Allow status updates for affiliates and operations
        allowed_updates = ['status', 'notes']
        if user.user_type == 'affiliate':
            # Affiliates can only update status and notes
            filtered_data = {k: v for k, v in request.data.items() if k in allowed_updates}
            request._full_data = filtered_data
        
        return super().update(request, *args, **kwargs)
    
    @action(detail=True, methods=['post'])
    def add_note(self, request, pk=None):
        """Add a note to a lead - with affiliate restrictions"""
        try:
            lead = self.get_object()
            user = request.user
            
            # Check if affiliate can only add notes to their own leads
            if user.user_type == 'affiliate':
                try:
                    affiliate = Affiliate.objects.get(user=user)
                    if lead.affiliate != affiliate:
                        return Response(
                            {'error': 'You can only add notes to your own leads'}, 
                            status=status.HTTP_403_FORBIDDEN
                        )
                except Affiliate.DoesNotExist:
                    return Response(
                        {'error': 'Affiliate profile not found'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            note_text = request.data.get('note', '').strip()
            
            if not note_text:
                return Response({'error': 'Note content is required'}, status=400)
            
            note = LeadNote.objects.create(
                lead=lead,
                user=request.user,
                note=note_text
            )
            
            return Response(LeadNoteSerializer(note).data, status=201)
        except Exception as e:
            logger.error(f"Error adding note: {e}")
            return Response({'error': str(e)}, status=500)
    
    @action(detail=True, methods=['get'])
    def notes(self, request, pk=None):
        """Get all notes for a lead - with affiliate restrictions"""
        try:
            lead = self.get_object()
            user = request.user
            
            # Check if affiliate can only view notes for their own leads
            if user.user_type == 'affiliate':
                try:
                    affiliate = Affiliate.objects.get(user=user)
                    if lead.affiliate != affiliate:
                        return Response(
                            {'error': 'You can only view notes for your own leads'}, 
                            status=status.HTTP_403_FORBIDDEN
                        )
                except Affiliate.DoesNotExist:
                    return Response(
                        {'error': 'Affiliate profile not found'}, 
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            notes = lead.lead_notes.all().order_by('-created_at')
            return Response(LeadNoteSerializer(notes, many=True).data)
        except Exception as e:
            logger.error(f"Error getting notes: {e}")
            return Response({'error': str(e)}, status=500)

class ExportLeadsView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            # Get filtered leads based on query parameters and user role
            user = request.user
            queryset = Lead.objects.select_related('form', 'affiliate')
            
            # Apply role-based filtering
            if user.user_type == 'affiliate':
                try:
                    affiliate = Affiliate.objects.get(user=user)
                    queryset = queryset.filter(affiliate=affiliate)
                except Affiliate.DoesNotExist:
                    queryset = Lead.objects.none()
            
            # Apply search and filters
            search = request.query_params.get('search')
            if search:
                queryset = queryset.filter(
                    Q(email__icontains=search) |
                    Q(name__icontains=search)
                )
            
            status_filter = request.query_params.get('status')
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            
            utm_source = request.query_params.get('utm_source')
            if utm_source:
                queryset = queryset.filter(utm_source=utm_source)
            
            form_id = request.query_params.get('form')
            if form_id:
                queryset = queryset.filter(form__id=form_id)
            
            # Create Excel workbook using openpyxl (no pandas needed)
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads Export"
            
            # Define headers
            headers = [
                'Email', 'Name', 'Phone', 'Form', 'Status', 'Affiliate',
                'UTM Source', 'UTM Medium', 'UTM Campaign', 'Created', 'Updated', 'IP Address'
            ]
            
            # Style the header row
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
            
            # Add data rows
            for row_num, lead in enumerate(queryset.order_by('-created_at'), 2):
                ws.cell(row=row_num, column=1, value=lead.email)
                ws.cell(row=row_num, column=2, value=lead.name)
                ws.cell(row=row_num, column=3, value=lead.phone)
                ws.cell(row=row_num, column=4, value=lead.form.name if lead.form else '')
                ws.cell(row=row_num, column=5, value=lead.get_status_display())
                ws.cell(row=row_num, column=6, value=lead.affiliate_code)
                ws.cell(row=row_num, column=7, value=lead.utm_source)
                ws.cell(row=row_num, column=8, value=lead.utm_medium)
                ws.cell(row=row_num, column=9, value=lead.utm_campaign)
                ws.cell(row=row_num, column=10, value=lead.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=11, value=lead.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=12, value=lead.ip_address)
                
                # Add form data columns dynamically
                form_data = lead.form_data or {}
                for key, value in form_data.items():
                    if key not in ['email', 'name', 'phone']:  # Avoid duplicates
                        # Find or create column for this form field
                        header_key = f'Form_{key.title()}'
                        if header_key not in headers:
                            headers.append(header_key)
                            col_num = len(headers)
                            # Add header if it's the first time we see this field
                            if row_num == 2:  # First data row
                                cell = ws.cell(row=1, column=col_num)
                                cell.value = header_key
                                cell.font = header_font
                                cell.fill = header_fill
                        
                        col_num = headers.index(header_key) + 1
                        ws.cell(row=row_num, column=col_num, value=str(value) if value else '')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Include affiliate code in filename if user is affiliate
            filename_suffix = ""
            if user.user_type == 'affiliate':
                try:
                    affiliate = Affiliate.objects.get(user=user)
                    filename_suffix = f"_{affiliate.affiliate_code}"
                except:
                    pass
            
            filename = f"leads_export{filename_suffix}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Save workbook to response
            wb.save(response)
            
            return response
            
        except Exception as e:
            logger.error(f"Export error: {e}")
            return Response({'error': 'Export failed'}, status=500)

class LeadStatsView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            user = request.user
            queryset = Lead.objects.all()
            
            # Apply role-based filtering
            if user.user_type == 'affiliate':
                try:
                    affiliate = Affiliate.objects.get(user=user)
                    queryset = queryset.filter(affiliate=affiliate)
                except Affiliate.DoesNotExist:
                    queryset = Lead.objects.none()
            
            # Calculate statistics
            total_leads = queryset.count()
            new_leads = queryset.filter(status='new').count()
            qualified_leads = queryset.filter(
                status__in=['qualified', 'demo_scheduled', 'demo_completed']
            ).count()
            closed_won = queryset.filter(status='closed_won').count()
            closed_lost = queryset.filter(status='closed_lost').count()
            
            # Conversion rates
            qualification_rate = (qualified_leads / total_leads * 100) if total_leads > 0 else 0
            close_rate = (closed_won / total_leads * 100) if total_leads > 0 else 0
            
            # Recent trends (last 30 days vs previous 30 days)
            now = timezone.now()
            thirty_days_ago = now - timedelta(days=30)
            sixty_days_ago = now - timedelta(days=60)
            
            recent_leads = queryset.filter(created_at__gte=thirty_days_ago).count()
            previous_leads = queryset.filter(
                created_at__gte=sixty_days_ago,
                created_at__lt=thirty_days_ago
            ).count()
            
            growth_rate = 0
            if previous_leads > 0:
                growth_rate = ((recent_leads - previous_leads) / previous_leads * 100)
            
            # Top sources
            top_sources = queryset.values('utm_source').annotate(
                count=Count('id')
            ).order_by('-count')[:10]
            
            # Form performance (only if user has access to multiple forms)
            form_performance = []
            if user.user_type in ['admin', 'operations']:
                form_performance = queryset.values('form__name').annotate(
                    count=Count('id')
                ).order_by('-count')[:10]
            elif user.user_type == 'affiliate':
                # Show form performance for affiliate's assigned forms
                try:
                    affiliate = Affiliate.objects.get(user=user)
                    form_performance = queryset.filter(
                        form__in=affiliate.assigned_forms.all()
                    ).values('form__name').annotate(
                        count=Count('id')
                    ).order_by('-count')[:10]
                except:
                    pass
            
            # Daily submissions for the last 30 days
            daily_data = []
            for i in range(30):
                date = (now - timedelta(days=i)).date()
                count = queryset.filter(created_at__date=date).count()
                daily_data.append({
                    'date': date.strftime('%Y-%m-%d'),
                    'submissions': count
                })
            
            response_data = {
                'total_leads': total_leads,
                'new_leads': new_leads,
                'qualified_leads': qualified_leads,
                'closed_won': closed_won,
                'closed_lost': closed_lost,
                'qualification_rate': round(qualification_rate, 2),
                'close_rate': round(close_rate, 2),
                'recent_leads': recent_leads,
                'growth_rate': round(growth_rate, 2),
                'top_sources': list(top_sources),
                'form_performance': list(form_performance),
                'daily_data': daily_data[::-1],  # Reverse for chronological order
                'user_type': user.user_type
            }
            
            # Add affiliate-specific data
            if user.user_type == 'affiliate':
                try:
                    affiliate = Affiliate.objects.get(user=user)
                    response_data.update({
                        'affiliate_code': affiliate.affiliate_code,
                        'assigned_forms_count': affiliate.assigned_forms.filter(is_active=True).count(),
                        'conversion_rate': affiliate.conversion_rate
                    })
                except:
                    pass
            
            return Response(response_data)
            
        except Exception as e:
            logger.error(f"Stats error: {e}")
            return Response({'error': str(e)}, status=500)
